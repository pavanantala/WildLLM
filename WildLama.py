#!/usr/bin/env python3
"""
MCC-Focused Hyperparameter Tuning for Llama-3.1-8B with LoRA
==============================================================

This script performs comprehensive hyperparameter tuning for tweet classification
using Llama-3.1-8B with LoRA (Low-Rank Adaptation). The primary evaluation metric
is MCC (Matthews Correlation Coefficient).

Features:
- Epoch-wise evaluation on train, validation, and test sets
- Comprehensive error analysis and classification pattern detection
- Excel export with detailed metrics and visualizations
- MCC-focused model selection and analysis
"""

import os
import sys
import argparse
import json
import random
import logging
import itertools
import shutil
from datetime import datetime
from collections import Counter

import numpy as np
import pandas as pd
import torch
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader

# Optional imports with fallbacks
try:
    import yaml
except ImportError:
    print("Warning: yaml package not found. Using alternative implementation.")
    class SimpleYAML:
        @staticmethod
        def dump(data, file):
            for key, value in data.items():
                file.write(f"{key}: {value}\n")
    yaml = SimpleYAML()

try:
    from transformers import AutoTokenizer, AutoModelForSequenceClassification
except ImportError:
    print("Warning: transformers package not found. Please install with: pip install transformers")
    sys.exit(1)
    
try:
    from peft import get_peft_model, LoraConfig, TaskType
except ImportError:
    print("Warning: peft package not found. Please install with: pip install peft")
    sys.exit(1)
    
try:
    from sklearn.metrics import (
        classification_report, confusion_matrix, roc_curve, auc, 
        precision_recall_curve, average_precision_score, roc_auc_score,
        matthews_corrcoef, precision_recall_fscore_support, accuracy_score
    )
except ImportError:
    print("Warning: scikit-learn package not found. Please install with: pip install scikit-learn")
    sys.exit(1)
    
try:
    import matplotlib.pyplot as plt
    import seaborn as sns
except ImportError:
    print("Warning: matplotlib/seaborn not found. Please install with: pip install matplotlib seaborn")
    
try:
    import openpyxl
    from openpyxl.chart import LineChart, Reference
except ImportError:
    print("Warning: openpyxl package not found. Please install with: pip install openpyxl")
    
try:
    from tqdm import tqdm
except ImportError:
    print("Warning: tqdm package not found. Using simple progress indicator instead.")
    def tqdm(iterable, **kwargs):
        desc = kwargs.get('desc', '')
        total = len(iterable) if hasattr(iterable, '__len__') else kwargs.get('total', None)
        print(f"Starting {desc}...")
        for i, item in enumerate(iterable):
            if i % 10 == 0 and total:
                print(f"{desc}: {i}/{total} ({i/total*100:.1f}%)")
            yield item
        print(f"Finished {desc}!")


# ================================================================================
# CONFIGURATION AND CONSTANTS
# ================================================================================

# Classification prompt for the model
CLASSIFICATION_PROMPT = """Classify if this tweet indicates wildlife trafficking or illegal ivory trade. Positive: selling, buying, pricing, auctioning, or trading ivory or protected animal products. Negative: conservation news, history, or ivory as color reference."""

# Base configuration
BASE_CONFIG = {
    'seed': 1,
    'num_workers': 2,
    'cv': 1,
    'rerun': 1,
    'use_cuda': True,
    'clip_grad': 10.0,
    'gamma': 0.95,
    'use_desc': False,
    'use_ocr': False,
    'use_synthetic': False,
    'weighted_0': 1.0,
    'weighted_1': 10.0,
    'weighted_loss': True,
    'dataPath': "data/",
    'textPath': ".csv",
    'num_classes': 2,
    'resultPath': "result.npy",
    'lossPath': "loss.pdf",
    'configWritePath': "config.yml",
    'errorAnalysisPath': "error_analysis.txt",
    'hidden_dim': 512,
    'epoch': 8  
}

# Hyperparameter search space
HYPERPARAMS = {
    'lr': [1e-4],
    'batch_size': [8, 16, 32]
}


# ================================================================================
# UTILITY FUNCTIONS
# ================================================================================

def set_seed(seed):
    """Set seed for reproducibility across all random number generators."""
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False


def setup_logger(name, log_file, level=logging.INFO):
    """Set up a logger with both file and console output."""
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    
    handler = logging.FileHandler(log_file)
    handler.setFormatter(formatter)
    
    logger = logging.getLogger(name)
    logger.setLevel(level)
    logger.addHandler(handler)
    logger.addHandler(logging.StreamHandler())
    
    return logger


# ================================================================================
# DATASET CLASS
# ================================================================================

class TweetDataset(Dataset):
    """Dataset class for tweet classification with optional user description."""
    
    def __init__(self, dataframe, tokenizer, max_length=128, config=None):
        self.data = dataframe
        self.tokenizer = tokenizer
        self.max_length = max_length
        self.config = config if config is not None else {}
        
        prompt_tokens = len(tokenizer.encode(CLASSIFICATION_PROMPT))
        print(f"Classification prompt uses {prompt_tokens} tokens")

    def __len__(self):
        return len(self.data)
    
    def __getitem__(self, idx):
        tweet = self.data.iloc[idx]['tweet_text_cleaned']
        
        # Include user description if specified
        if (self.config.get('use_desc', False) and 
            'user_description_cleaned' in self.data.columns and 
            not pd.isna(self.data.iloc[idx]['user_description_cleaned'])):
            user_desc = self.data.iloc[idx]['user_description_cleaned']
            tweet_content = f"Tweet: {tweet} User description: {user_desc}"
        else:
            tweet_content = f"Tweet: {tweet}"
        
        # Use classification prompt
        text = f"{CLASSIFICATION_PROMPT} {tweet_content}"
            
        # Tokenize
        encoding = self.tokenizer(
            text,
            truncation=True,
            max_length=self.max_length,
            padding='max_length',
            return_tensors='pt'
        )
        
        return {
            'input_ids': encoding['input_ids'].squeeze(),
            'attention_mask': encoding['attention_mask'].squeeze(),
            'labels': torch.tensor(self.data.iloc[idx]['label'], dtype=torch.long)
        }


# ================================================================================
# MODEL TRAINING AND EVALUATION
# ================================================================================

def train_epoch(model, dataloader, optimizer, criterion, device, clip_grad=None):
    """Train model for one epoch and return metrics."""
    model.train()
    epoch_loss = 0
    all_preds, all_labels, all_probs = [], [], []
    
    for batch in tqdm(dataloader, desc="Training"):
        optimizer.zero_grad()
        
        input_ids = batch['input_ids'].to(device)
        attention_mask = batch['attention_mask'].to(device)
        labels = batch['labels'].to(device)
        
        outputs = model(input_ids=input_ids, attention_mask=attention_mask, labels=labels)
        loss = outputs.loss
        logits = outputs.logits
        
        loss.backward()
        
        if clip_grad is not None:
            torch.nn.utils.clip_grad_norm_(model.parameters(), clip_grad)
            
        optimizer.step()
        epoch_loss += loss.item()
        
        # Get predictions and probabilities
        probs = torch.softmax(logits, dim=1)
        preds = torch.argmax(logits, dim=1).cpu().numpy()
        
        all_preds.extend(preds)
        all_labels.extend(labels.cpu().numpy())
        all_probs.extend(probs.detach().cpu().float().numpy())
    
    return epoch_loss / len(dataloader), all_preds, all_labels, all_probs


def evaluate(model, dataloader, criterion, device):
    """Evaluate model and return metrics."""
    model.eval()
    epoch_loss = 0
    all_preds, all_labels, all_probs = [], [], []
    
    with torch.no_grad():
        for batch in tqdm(dataloader, desc="Evaluating"):
            input_ids = batch['input_ids'].to(device)
            attention_mask = batch['attention_mask'].to(device)
            labels = batch['labels'].to(device)
            
            outputs = model(input_ids=input_ids, attention_mask=attention_mask, labels=labels)
            loss = outputs.loss
            logits = outputs.logits
            
            epoch_loss += loss.item()
            
            # Get predictions and probabilities
            probs = torch.softmax(logits, dim=1)
            preds = torch.argmax(logits, dim=1).cpu().numpy()
            
            all_preds.extend(preds)
            all_labels.extend(labels.cpu().numpy())
            all_probs.extend(probs.detach().cpu().float().numpy())
    
    return epoch_loss / len(dataloader), all_preds, all_labels, all_probs


# ================================================================================
# METRICS CALCULATION
# ================================================================================

def calculate_metrics(true_labels, predicted_labels, predicted_probs=None):
    """Calculate comprehensive classification metrics with MCC as primary focus."""
    # Calculate MCC (Matthews Correlation Coefficient) - PRIMARY METRIC
    mcc = matthews_corrcoef(true_labels, predicted_labels)
    
    # Calculate precision, recall, F1 for each class
    precision, recall, f1, _ = precision_recall_fscore_support(
        true_labels, predicted_labels, average=None, zero_division=0
    )
    
    # Calculate macro average metrics
    macro_precision, macro_recall, macro_f1, _ = precision_recall_fscore_support(
        true_labels, predicted_labels, average='macro', zero_division=0
    )
    
    # Calculate accuracy
    accuracy = accuracy_score(true_labels, predicted_labels)
    
    # Calculate confusion matrix components
    cm = confusion_matrix(true_labels, predicted_labels)
    
    # For binary classification
    if len(np.unique(true_labels)) == 2:
        if cm.shape == (2, 2):
            tn, fp, fn, tp = cm.ravel()
        else:
            tn = fp = fn = tp = 0
            if cm.shape == (1, 1):
                if np.unique(predicted_labels)[0] == 0:
                    tn = cm[0, 0]
                else:
                    tp = cm[0, 0]
        
        # Calculate AUC if probabilities are provided
        auc_score = None
        if predicted_probs is not None:
            try:
                if isinstance(predicted_probs, list):
                    predicted_probs = np.array(predicted_probs)
                
                if predicted_probs.ndim == 2 and predicted_probs.shape[1] >= 2:
                    pos_probs = predicted_probs[:, 1]  
                else:
                    pos_probs = predicted_probs
                
                auc_score = roc_auc_score(true_labels, pos_probs)
            except Exception as e:
                print(f"Warning: Could not calculate AUC score: {e}")
                auc_score = None
        
        # Get metrics for positive class (class 1)
        pos_precision = precision[1] if len(precision) > 1 else (
            tp / (tp + fp) if (tp + fp) > 0 else 0
        )
        pos_recall = recall[1] if len(recall) > 1 else (
            tp / (tp + fn) if (tp + fn) > 0 else 0
        )
        pos_f1 = f1[1] if len(f1) > 1 else (
            2 * pos_precision * pos_recall / (pos_precision + pos_recall) 
            if (pos_precision + pos_recall) > 0 else 0
        )
        
    else:
        # Multi-class case
        tn = fp = fn = tp = None
        auc_score = None  
        pos_precision = None
        pos_recall = None
        pos_f1 = None
    
    return {
        'mcc': mcc,  
        'accuracy': accuracy,
        'macro_precision': macro_precision,
        'macro_recall': macro_recall, 
        'macro_f1': macro_f1,
        'pos_precision': pos_precision,
        'pos_recall': pos_recall,
        'pos_f1': pos_f1,
        'auc': auc_score,
        'confusion_matrix': cm.tolist(),
        'tn': int(tn) if tn is not None else None,
        'fp': int(fp) if fp is not None else None,
        'fn': int(fn) if fn is not None else None,
        'tp': int(tp) if tp is not None else None
    }


# ================================================================================
# ERROR ANALYSIS FUNCTIONS
# ================================================================================

def save_predictions_with_errors(all_preds, all_labels, all_probs, dataset_df, 
                                output_path, dataset_name, config=None):
    """Save predictions with enhanced data for error analysis."""
    
    # Create DataFrame with predictions and enhanced information
    results_df = dataset_df.copy()
    results_df['predicted_label'] = all_preds
    results_df['true_label'] = all_labels
    results_df['prob_class_0'] = [prob[0] for prob in all_probs]
    results_df['prob_class_1'] = [prob[1] for prob in all_probs]
    results_df['prediction_confidence'] = [max(prob) for prob in all_probs]
    
    # Add classification prompt and formatted text
    enhanced_texts = []
    for idx, row in results_df.iterrows():
        tweet = row['tweet_text_cleaned']
        if (config and config.get('use_desc', False) and 
            'user_description_cleaned' in results_df.columns and 
            not pd.isna(row['user_description_cleaned'])):
            user_desc = row['user_description_cleaned']
            tweet_content = f"Tweet: {tweet} User description: {user_desc}"
        else:
            tweet_content = f"Tweet: {tweet}"
        
        formatted_text = f"{CLASSIFICATION_PROMPT}\n\nText to classify: {tweet_content}\n\nLabel (0 or 1):"
        enhanced_texts.append(formatted_text)
    
    results_df['formatted_input_text'] = enhanced_texts
    results_df['original_tweet_only'] = results_df['tweet_text_cleaned']
    
    # Add classification type
    classification_types = []
    for true_label, pred_label in zip(all_labels, all_preds):
        if true_label == 1 and pred_label == 1:
            classification_types.append('True Positive')
        elif true_label == 0 and pred_label == 0:
            classification_types.append('True Negative')
        elif true_label == 0 and pred_label == 1:
            classification_types.append('False Positive')  
        elif true_label == 1 and pred_label == 0:
            classification_types.append('False Negative')  
    
    results_df['classification_type'] = classification_types
    results_df['is_correct'] = results_df['true_label'] == results_df['predicted_label']
    results_df['dataset_split'] = dataset_name
    
    # Save main results
    results_df.to_csv(output_path, index=False)
    
    # Create separate files for each classification type
    error_analysis_columns = [
        'original_tweet_only', 'formatted_input_text', 'true_label', 'predicted_label', 
        'prob_class_0', 'prob_class_1', 'prediction_confidence', 'classification_type', 
        'is_correct', 'dataset_split'
    ]
    
    # Include additional columns if they exist
    optional_columns = ['user_description_cleaned', 'tweet_id', 'user_id']
    for col in optional_columns:
        if col in results_df.columns:
            error_analysis_columns.insert(-4, col)
    
    # Split by classification type
    classification_dfs = {}
    for class_type in ['True Positive', 'True Negative', 'False Positive', 'False Negative']:
        df_subset = results_df[results_df['classification_type'] == class_type][error_analysis_columns]
        classification_dfs[class_type] = df_subset
        
        if len(df_subset) > 0:
            file_suffix = class_type.lower().replace(' ', '_') + '_for_analysis.csv'
            subset_path = output_path.replace('.csv', f'_{file_suffix}')
            df_subset.to_csv(subset_path, index=False)
            print(f"{class_type} for analysis saved to: {subset_path} ({len(df_subset)} instances)")
    
    # Save combined error file
    errors_df = pd.concat([
        classification_dfs['False Positive'], 
        classification_dfs['False Negative']
    ], ignore_index=True)
    
    if len(errors_df) > 0:
        errors_path = output_path.replace('.csv', '_all_errors_for_analysis.csv')
        errors_df.to_csv(errors_path, index=False)
        print(f"All errors for analysis saved to: {errors_path} ({len(errors_df)} instances)")
    
    return (results_df, classification_dfs['True Positive'], classification_dfs['True Negative'],
            classification_dfs['False Positive'], classification_dfs['False Negative'])


def analyze_classification_patterns(tp_df, tn_df, fp_df, fn_df, dataset_name):
    """Analyze patterns in all four classification types."""
    
    print(f"\n{dataset_name} Complete Classification Analysis:")
    print("="*60)
    
    total_instances = len(tp_df) + len(tn_df) + len(fp_df) + len(fn_df)
    correct_instances = len(tp_df) + len(tn_df)
    
    print(f"Total instances: {total_instances}")
    print(f"Correct predictions: {correct_instances} ({correct_instances/total_instances*100:.1f}%)")
    print(f"Incorrect predictions: {len(fp_df) + len(fn_df)} ({(len(fp_df) + len(fn_df))/total_instances*100:.1f}%)")
    
    # Use available text column
    text_column = 'original_tweet_only' if any(
        'original_tweet_only' in df.columns for df in [tp_df, tn_df, fp_df, fn_df] if len(df) > 0
    ) else 'tweet_text_cleaned'
    
    # Analyze each classification type
    analysis_data = [
        (tp_df, "True Positives", "correctly identified as positive", 'prob_class_1'),
        (tn_df, "True Negatives", "correctly identified as negative", 'prob_class_0'),
        (fp_df, "False Positives", "wrongly predicted as positive", 'prob_class_1'),
        (fn_df, "False Negatives", "wrongly predicted as negative", 'prob_class_0')
    ]
    
    for df, name, description, prob_col in analysis_data:
        if len(df) > 0:
            print(f"\n{name} Analysis ({len(df)} instances, {len(df)/total_instances*100:.1f}%):")
            print(f"Average confidence: {df[prob_col].mean():.4f}")
            
            df_copy = df.copy()
            df_copy['text_length'] = df_copy[text_column].str.len()
            print(f"Text length - Mean: {df_copy['text_length'].mean():.1f}, "
                  f"Median: {df_copy['text_length'].median():.1f}")
            
            print(f"Sample {name} ({description}):")
            for i, row in df.head(2).iterrows():
                print(f"  - Text: '{row[text_column][:80]}...'")
                print(f"    Confidence: {row[prob_col]:.4f}")


# ================================================================================
# EXCEL ANALYSIS FUNCTIONS
# ================================================================================

def create_excel_mcc_analysis(all_results, output_dir):
    """Create Excel file with epoch-wise MCC analysis and graphs including test results."""
    excel_path = os.path.join(output_dir, 'mcc_epoch_wise_analysis.xlsx')
    
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Sheet 1: Epoch-wise Data
        ws_data = wb.create_sheet("Epoch-wise MCC Data")
        
        headers = ['Experiment_ID', 'Learning_Rate', 'Batch_Size', 'Epoch', 
                   'Train_MCC', 'Val_MCC', 'Test_MCC',
                   'Train_Accuracy', 'Val_Accuracy', 'Test_Accuracy',
                   'Train_F1', 'Val_F1', 'Test_F1',
                   'Train_AUC', 'Val_AUC', 'Test_AUC']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws_data.cell(row=1, column=col, value=header)
        
        # Fill data
        row_idx = 2
        experiment_id = 1
        
        for result in all_results:
            lr = result['hyperparameters']['learning_rate']
            bs = result['hyperparameters']['batch_size']
            epochs = result['hyperparameters']['epochs']
            
            # Get epoch-wise data
            train_mcc_hist = result['epoch_wise_data']['train_mcc_history']
            val_mcc_hist = result['epoch_wise_data']['val_mcc_history']
            test_mcc_hist = result['epoch_wise_data']['test_mcc_history']
            
            train_metrics_hist = result['epoch_wise_data']['train_metrics_history']
            val_metrics_hist = result['epoch_wise_data']['val_metrics_history']
            test_metrics_hist = result['epoch_wise_data']['test_metrics_history']
            
            # Write data for each epoch
            for epoch in range(epochs):
                data_row = [
                    f"Exp_{experiment_id}", lr, bs, epoch + 1,
                    round(train_mcc_hist[epoch], 4),
                    round(val_mcc_hist[epoch], 4),
                    round(test_mcc_hist[epoch], 4),
                    round(train_metrics_hist[epoch]['accuracy'], 4),
                    round(val_metrics_hist[epoch]['accuracy'], 4),
                    round(test_metrics_hist[epoch]['accuracy'], 4),
                    round(train_metrics_hist[epoch]['pos_f1'], 4),
                    round(val_metrics_hist[epoch]['pos_f1'], 4),
                    round(test_metrics_hist[epoch]['pos_f1'], 4),
                    round(train_metrics_hist[epoch]['auc'], 4) if train_metrics_hist[epoch]['auc'] else 0,
                    round(val_metrics_hist[epoch]['auc'], 4) if val_metrics_hist[epoch]['auc'] else 0,
                    round(test_metrics_hist[epoch]['auc'], 4) if test_metrics_hist[epoch]['auc'] else 0
                ]
                
                for col, value in enumerate(data_row, 1):
                    ws_data.cell(row=row_idx, column=col, value=value)
                
                row_idx += 1
            
            experiment_id += 1
        
        # Auto-adjust column widths
        for column in ws_data.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws_data.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 2: Best Performance Analysis
        ws_best = wb.create_sheet("Best Performance Analysis")
        
        # Find best performing configurations
        best_by_test_mcc = max(all_results, key=lambda x: max(x['epoch_wise_data']['test_mcc_history']))
        best_by_val_mcc = max(all_results, key=lambda x: max(x['epoch_wise_data']['val_mcc_history']))
        
        # Create analysis table
        analysis_headers = ['Metric', 'Best_Learning_Rate', 'Best_Batch_Size', 'Best_Value', 'Best_Epoch', 'Dataset']
        
        for col, header in enumerate(analysis_headers, 1):
            ws_best.cell(row=1, column=col, value=header)
        
        # Test MCC best
        best_test_mcc_val = max(best_by_test_mcc['epoch_wise_data']['test_mcc_history'])
        best_test_mcc_epoch = best_by_test_mcc['epoch_wise_data']['test_mcc_history'].index(best_test_mcc_val) + 1
        
        best_data = [
            ["MCC", best_by_test_mcc['hyperparameters']['learning_rate'], 
             best_by_test_mcc['hyperparameters']['batch_size'], 
             round(best_test_mcc_val, 4), best_test_mcc_epoch, "TEST"],
            ["MCC", best_by_val_mcc['hyperparameters']['learning_rate'], 
             best_by_val_mcc['hyperparameters']['batch_size'], 
             round(max(best_by_val_mcc['epoch_wise_data']['val_mcc_history']), 4),
             best_by_val_mcc['epoch_wise_data']['val_mcc_history'].index(
                 max(best_by_val_mcc['epoch_wise_data']['val_mcc_history'])) + 1, "VALIDATION"]
        ]
        
        for row_idx, row_data in enumerate(best_data, 2):
            for col, value in enumerate(row_data, 1):
                ws_best.cell(row=row_idx, column=col, value=value)
        
        # Hyperparameter vs MCC analysis
        ws_best.cell(row=5, column=1, value="Learning Rate vs MCC Analysis")
        summary_headers = ['Learning_Rate', 'Batch_Size', 'Best_Train_MCC', 'Best_Val_MCC', 'Best_Test_MCC', 'Rank_by_Test_MCC']
        
        for col, header in enumerate(summary_headers, 1):
            ws_best.cell(row=6, column=col, value=header)
        
        # Sort results by best test MCC
        sorted_results = sorted(all_results, key=lambda x: max(x['epoch_wise_data']['test_mcc_history']), reverse=True)
        
        for idx, result in enumerate(sorted_results):
            row_num = 7 + idx
            summary_data = [
                result['hyperparameters']['learning_rate'],
                result['hyperparameters']['batch_size'],
                round(max(result['epoch_wise_data']['train_mcc_history']), 4),
                round(max(result['epoch_wise_data']['val_mcc_history']), 4),
                round(max(result['epoch_wise_data']['test_mcc_history']), 4),
                idx + 1
            ]
            
            for col, value in enumerate(summary_data, 1):
                ws_best.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths for best performance sheet
        for column in ws_best.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws_best.column_dimensions[column_letter].width = adjusted_width
        
        # Save the Excel file
        wb.save(excel_path)
        print(f"Excel file with epoch-wise MCC analysis saved to: {excel_path}")
        return excel_path
        
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        return None


# ================================================================================
# VISUALIZATION FUNCTIONS
# ================================================================================

def plot_comparative_results(all_results, output_dir):
    """Create comprehensive visualizations focused on MCC."""
    comparison_dir = os.path.join(output_dir, 'comparison')
    os.makedirs(comparison_dir, exist_ok=True)
    
    # Extract results for plotting
    configs = []
    mcc_scores = []
    accuracies = []
    f1_scores = []
    val_mccs = []
    auc_scores = []
    
    for result in all_results:
        hp = result['hyperparameters']
        config_name = f"lr={hp['learning_rate']}, bs={hp['batch_size']}"
        configs.append(config_name)
        mcc_scores.append(result['test_metrics']['mcc'])
        accuracies.append(result['test_metrics']['accuracy'])
        f1_scores.append(result['test_metrics']['pos_f1'])
        val_mccs.append(result['best_val_mcc'])
        auc_scores.append(result['test_metrics']['auc'] if result['test_metrics']['auc'] else 0)
    
    # Sort by MCC score
    sorted_indices = np.argsort(mcc_scores)[::-1]
    configs = [configs[i] for i in sorted_indices]
    mcc_scores = [mcc_scores[i] for i in sorted_indices]
    accuracies = [accuracies[i] for i in sorted_indices]
    f1_scores = [f1_scores[i] for i in sorted_indices]
    val_mccs = [val_mccs[i] for i in sorted_indices]
    auc_scores = [auc_scores[i] for i in sorted_indices]
    
    # Primary MCC comparison chart
    plt.figure(figsize=(14, 8))
    bars = plt.bar(configs, mcc_scores, color='darkblue')
    plt.xlabel('Hyperparameters')
    plt.ylabel('MCC Score')
    plt.title('MCC Score Comparison Across Different Hyperparameters (PRIMARY METRIC)')
    plt.xticks(rotation=45, ha='right')
    plt.ylim(-1.0, 1.0)
    
    # Add value labels
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height + 0.02,
                f'{height:.3f}', ha='center', va='bottom', rotation=0, fontweight='bold')
    
    plt.axhline(y=0, color='red', linestyle='--', alpha=0.5, label='Random Performance (MCC=0)')
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(comparison_dir, 'mcc_comparison_primary.pdf'))
    plt.savefig(os.path.join(comparison_dir, 'mcc_comparison_primary.png'), dpi=300)
    plt.close()
    
    # Multiple metrics comparison
    plt.figure(figsize=(16, 10))
    x = np.arange(len(configs))
    width = 0.15
    
    plt.bar(x - width*2, mcc_scores, width, label='MCC (Primary)', color='darkblue', edgecolor='black', linewidth=1.5)
    plt.bar(x - width, accuracies, width, label='Accuracy', color='skyblue')
    plt.bar(x, f1_scores, width, label='F1 Score', color='orange')
    plt.bar(x + width, auc_scores, width, label='AUC', color='purple')
    plt.bar(x + width*2, val_mccs, width, label='Val MCC', color='green')
    
    plt.xlabel('Hyperparameters')
    plt.ylabel('Score')
    plt.title('Comprehensive Performance Metrics Comparison (MCC as Primary)')
    plt.xticks(x, configs, rotation=45, ha='right')
    plt.ylim(-1.0, 1.0)
    plt.legend()
    plt.axhline(y=0, color='gray', linestyle='--', alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(comparison_dir, 'comprehensive_metrics_mcc_focused.pdf'))
    plt.close()
    
    # Best model performance
    best_model = max(all_results, key=lambda x: x['test_metrics']['mcc'])
    best_config = f"lr={best_model['hyperparameters']['learning_rate']}, bs={best_model['hyperparameters']['batch_size']}"
    
    plt.figure(figsize=(12, 8))
    metrics = ['MCC (Primary)', 'Accuracy', 'F1 Score', 'AUC']
    best_values = [
        best_model['test_metrics']['mcc'],
        best_model['test_metrics']['accuracy'],
        best_model['test_metrics']['pos_f1'],
        best_model['test_metrics']['auc'] if best_model['test_metrics']['auc'] else 0
    ]
    
    colors = ['darkblue', 'skyblue', 'orange', 'purple']
    bars = plt.bar(metrics, best_values, color=colors)
    
    # Emphasize MCC bar
    bars[0].set_edgecolor('black')
    bars[0].set_linewidth(3)
    
    for bar, value in zip(bars, best_values):
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height + 0.02,
                f'{value:.3f}', ha='center', va='bottom', 
                fontweight='bold' if bar == bars[0] else 'normal')
    
    plt.ylim(-1.0, 1.0)
    plt.title(f'Best Model Performance (Selected by MCC): {best_config}')
    plt.ylabel('Score')
    plt.axhline(y=0, color='gray', linestyle='--', alpha=0.3)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig(os.path.join(comparison_dir, 'best_model_mcc_focused.pdf'))
    plt.close()
    
    return best_model


# ================================================================================
# MAIN EXPERIMENT FUNCTION
# ================================================================================

def run_experiment(config, train_df, val_df, test_df):
    """Run a single experiment with given hyperparameters."""
    # Create experiment directory
    experiment_dir = os.path.join(
        config['outputPath'], 
        f"lr_{config['lr']}_bs_{config['batch_size']}"
    )
    os.makedirs(experiment_dir, exist_ok=True)
    
    # Setup logging
    log_file = os.path.join(experiment_dir, "log.txt")
    logger = setup_logger(f"exp_lr_{config['lr']}_bs_{config['batch_size']}", log_file)
    
    # Save config
    config_path = os.path.join(experiment_dir, config['configWritePath'])
    with open(config_path, 'w') as f:
        yaml.dump(config, f)
    
    # Set seed and device
    set_seed(config['seed'])
    device = torch.device("cuda" if torch.cuda.is_available() and config['use_cuda'] else "cpu")
    logger.info(f"Using device: {device}")
    
    # Initialize tokenizer and model
    tokenizer = AutoTokenizer.from_pretrained("meta-llama/Llama-3.1-8B")
    if tokenizer.pad_token is None:
        tokenizer.pad_token = tokenizer.eos_token
        model_config = {'pad_token_id': tokenizer.eos_token_id}
    else:
        model_config = {}
    
    # Create datasets and dataloaders
    train_dataset = TweetDataset(train_df, tokenizer, config=config)
    val_dataset = TweetDataset(val_df, tokenizer, config=config)
    test_dataset = TweetDataset(test_df, tokenizer, config=config)
    
    train_loader = DataLoader(train_dataset, batch_size=config['batch_size'], 
                            shuffle=True, num_workers=config['num_workers'])
    val_loader = DataLoader(val_dataset, batch_size=config['batch_size'], 
                          shuffle=False, num_workers=config['num_workers'])
    test_loader = DataLoader(test_dataset, batch_size=config['batch_size'], 
                           shuffle=False, num_workers=config['num_workers'])
    
    # Set up model with LoRA
    peft_config = LoraConfig(
        task_type=TaskType.SEQ_CLS,
        inference_mode=False,
        r=16,
        lora_alpha=32,
        lora_dropout=0.1,
        target_modules=["q_proj", "v_proj", "k_proj", "o_proj", "gate_proj", "up_proj", "down_proj"]
    )
    
    model = AutoModelForSequenceClassification.from_pretrained(
        "meta-llama/Llama-3.1-8B",
        num_labels=config['num_classes'],
        torch_dtype=torch.bfloat16 if torch.cuda.is_available() else torch.float32,
        **model_config
    )
    
    model = get_peft_model(model, peft_config)
    trainable_params = sum(p.numel() for p in model.parameters() if p.requires_grad)
    total_params = sum(p.numel() for p in model.parameters())
    logger.info(f"Trainable parameters: {trainable_params} ({trainable_params/total_params*100:.2f}% of total)")
    
    model = model.to(device)
    
    # Set up loss function and optimizer
    if config['weighted_loss']:
        weights = torch.tensor([config['weighted_0'], config['weighted_1']], dtype=torch.float32).to(device)
        criterion = nn.CrossEntropyLoss(weight=weights)
    else:
        criterion = nn.CrossEntropyLoss()
    
    optimizer = torch.optim.AdamW(model.parameters(), lr=config['lr'])
    scheduler = torch.optim.lr_scheduler.ExponentialLR(optimizer, gamma=config['gamma'])
    
    # Training loop with epoch-wise evaluation
    train_losses, val_losses, test_losses = [], [], []
    train_mcc_history, val_mcc_history, test_mcc_history = [], [], []
    train_metrics_history, val_metrics_history, test_metrics_history = [], [], []
    
    best_val_mcc = -1.0
    best_val_metrics = None
    best_model_path = os.path.join(experiment_dir, "best_model.pkl")
    
    logger.info("Starting training with epoch-wise test evaluation...")
    
    for epoch in range(config['epoch']):
        logger.info(f"Epoch {epoch+1}/{config['epoch']}")
        
        # Train
        train_loss, train_preds, train_labels, train_probs = train_epoch(
            model, train_loader, optimizer, criterion, device, config['clip_grad']
        )
        
        # Validate
        val_loss, val_preds, val_labels, val_probs = evaluate(
            model, val_loader, criterion, device
        )
        
        # Test
        test_loss, test_preds, test_labels, test_probs = evaluate(
            model, test_loader, criterion, device
        )
        
        scheduler.step()
        
        # Calculate metrics
        train_metrics = calculate_metrics(train_labels, train_preds, train_probs)
        val_metrics = calculate_metrics(val_labels, val_preds, val_probs)
        test_metrics = calculate_metrics(test_labels, test_preds, test_probs)
        
        # Store metrics
        train_losses.append(train_loss)
        val_losses.append(val_loss)
        test_losses.append(test_loss)
        
        train_mcc_history.append(train_metrics['mcc'])
        val_mcc_history.append(val_metrics['mcc'])
        test_mcc_history.append(test_metrics['mcc'])
        
        train_metrics_history.append(train_metrics)
        val_metrics_history.append(val_metrics)
        test_metrics_history.append(test_metrics)
        
        # Log progress
        logger.info(f"Train MCC: {train_metrics['mcc']:.4f}, Loss: {train_loss:.4f}")
        logger.info(f"Val   MCC: {val_metrics['mcc']:.4f}, Loss: {val_loss:.4f}")
        logger.info(f"Test  MCC: {test_metrics['mcc']:.4f}, Loss: {test_loss:.4f}")
        
        # Save best model based on validation MCC
        if val_metrics['mcc'] > best_val_mcc:
            best_val_mcc = val_metrics['mcc']
            best_val_metrics = val_metrics
            torch.save(model.state_dict(), best_model_path)
            logger.info(f"New best model saved! (Val MCC: {best_val_mcc:.4f})")
    
    # Plot loss and MCC curves
    plt.figure(figsize=(12, 8))
    plt.plot(train_losses, label='Train Loss')
    plt.plot(val_losses, label='Validation Loss')
    plt.plot(test_losses, label='Test Loss')
    plt.xlabel('Epoch')
    plt.ylabel('Loss')
    plt.title(f'Loss Curves (lr={config["lr"]}, bs={config["batch_size"]})')
    plt.legend()
    plt.savefig(os.path.join(experiment_dir, config['lossPath']))
    plt.close()
    
    plt.figure(figsize=(12, 8))
    plt.plot(train_mcc_history, label='Train MCC', marker='o')
    plt.plot(val_mcc_history, label='Validation MCC', marker='s')
    plt.plot(test_mcc_history, label='Test MCC', marker='^')
    plt.xlabel('Epoch')
    plt.ylabel('MCC Score')
    plt.title(f'MCC Progression (lr={config["lr"]}, bs={config["batch_size"]})')
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.axhline(y=0, color='red', linestyle='--', alpha=0.5, label='Random Performance')
    plt.savefig(os.path.join(experiment_dir, 'mcc_progression.pdf'))
    plt.close()
    
    # Final evaluation with best model
    model.load_state_dict(torch.load(best_model_path))
    logger.info("Evaluating best model on test set...")
    
    final_test_loss, final_test_preds, final_test_labels, final_test_probs = evaluate(
        model, test_loader, criterion, device
    )
    final_val_loss, final_val_preds, final_val_labels, final_val_probs = evaluate(
        model, val_loader, criterion, device
    )
    
    final_test_metrics = calculate_metrics(final_test_labels, final_test_preds, final_test_probs)
    
    # Save predictions and perform error analysis
    test_predictions_path = os.path.join(experiment_dir, 'test_predictions_complete_analysis.csv')
    val_predictions_path = os.path.join(experiment_dir, 'val_predictions_complete_analysis.csv')
    
    test_results_df, test_tp_df, test_tn_df, test_fp_df, test_fn_df = save_predictions_with_errors(
        final_test_preds, final_test_labels, final_test_probs, 
        test_df, test_predictions_path, 'test', config
    )
    
    val_results_df, val_tp_df, val_tn_df, val_fp_df, val_fn_df = save_predictions_with_errors(
        final_val_preds, final_val_labels, final_val_probs,
        val_df, val_predictions_path, 'validation', config
    )
    
    # Analyze classification patterns
    analyze_classification_patterns(test_tp_df, test_tn_df, test_fp_df, test_fn_df, 'Test')
    analyze_classification_patterns(val_tp_df, val_tn_df, val_fp_df, val_fn_df, 'Validation')
    
    # Log final results
    auc_str = f"{final_test_metrics['auc']:.4f}" if final_test_metrics['auc'] else 'N/A'
    logger.info(f"Final Test MCC: {final_test_metrics['mcc']:.4f} (PRIMARY METRIC)")
    logger.info(f"Final Test Accuracy: {final_test_metrics['accuracy']:.4f}")
    logger.info(f"Final Test F1 Score: {final_test_metrics['pos_f1']:.4f}")
    logger.info(f"Final Test AUC: {auc_str}")
    
    # Prepare results
    results = {
        'hyperparameters': {
            'learning_rate': config['lr'],
            'batch_size': config['batch_size'],
            'epochs': config['epoch']
        },
        'best_val_mcc': best_val_mcc,
        'best_val_metrics': best_val_metrics,
        'test_loss': final_test_loss,
        'test_metrics': final_test_metrics,
        'train_losses': train_losses,
        'val_losses': val_losses,
        'test_losses': test_losses,
        'epoch_wise_data': {
            'train_mcc_history': train_mcc_history,
            'val_mcc_history': val_mcc_history,
            'test_mcc_history': test_mcc_history,
            'train_metrics_history': train_metrics_history,
            'val_metrics_history': val_metrics_history,
            'test_metrics_history': test_metrics_history
        }
    }
    
    with open(os.path.join(experiment_dir, 'results.json'), 'w') as f:
        json.dump(results, f, indent=4, default=lambda x: float(x) if isinstance(x, np.float32) else x)
    
    return results


# ================================================================================
# MAIN FUNCTION
# ================================================================================

def main():
    parser = argparse.ArgumentParser(description='Run MCC-focused hyperparameter tuning for Llama LoRA')
    parser.add_argument('--data_path', type=str, default='data/',
                        help='Path to data directory containing train.csv, val.csv, test.csv')
    parser.add_argument('--output_dir', type=str, default=None,
                        help='Output directory (default: LLM/llm_tuning_TIMESTAMP)')
    parser.add_argument('--subset', action='store_true',
                        help='Run only a subset of hyperparameters for testing')
    parser.add_argument('--no_cuda', action='store_true',
                        help='Disable CUDA (use CPU only)')
    args = parser.parse_args()
    
    # Set up main logger
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler("LLM/mcc_focused_llm_tuning.log"),
            logging.StreamHandler()
        ]
    )
    main_logger = logging.getLogger(__name__)
    
    # Create output directory
    if args.output_dir is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = f"LLM/llm_tuning_{timestamp}"
    else:
        output_dir = args.output_dir
    
    os.makedirs(output_dir, exist_ok=True)
    
    # Update config
    config = BASE_CONFIG.copy()
    config['dataPath'] = args.data_path
    config['outputPath'] = output_dir
    config['use_cuda'] = not args.no_cuda
    
    # Load data
    main_logger.info(f"Loading data from {config['dataPath']}")
    try:
        train_df = pd.read_csv(os.path.join(config['dataPath'], 'train.csv'))
        test_df = pd.read_csv(os.path.join(config['dataPath'], 'test.csv'))
        val_df = pd.read_csv(os.path.join(config['dataPath'], 'val.csv'))
    except FileNotFoundError as e:
        main_logger.error(f"Error loading data: {e}")
        print(f"Error: Could not find required CSV files in {config['dataPath']}")
        sys.exit(1)
    
    main_logger.info(f"Train: {train_df.shape}, Val: {val_df.shape}, Test: {test_df.shape}")
    
    # Set up hyperparameter combinations
    if args.subset:
        lr_values = [HYPERPARAMS['lr'][0], HYPERPARAMS['lr'][4], HYPERPARAMS['lr'][-1]]
        bs_values = [HYPERPARAMS['batch_size'][0], HYPERPARAMS['batch_size'][-1]]
        main_logger.info("Running with subset of hyperparameters for testing")
    else:
        lr_values = HYPERPARAMS['lr']
        bs_values = HYPERPARAMS['batch_size']
    
    param_combinations = list(itertools.product(lr_values, bs_values))
    main_logger.info(f"Running {len(param_combinations)} experiments")
    main_logger.info("PRIMARY EVALUATION METRIC: MCC (Matthews Correlation Coefficient)")
    
    # Run experiments
    all_results = []
    
    for lr, batch_size in param_combinations:
        run_config = config.copy()
        run_config['lr'] = lr
        run_config['batch_size'] = batch_size
        
        main_logger.info(f"Starting experiment with lr={lr}, batch_size={batch_size}")
        
        result = run_experiment(run_config, train_df, val_df, test_df)
        all_results.append(result)
        
        auc_str = f"{result['test_metrics']['auc']:.4f}" if result['test_metrics']['auc'] else 'N/A'
        main_logger.info(f"Completed: lr={lr}, bs={batch_size} - "
                        f"Test MCC: {result['test_metrics']['mcc']:.4f}, "
                        f"F1: {result['test_metrics']['pos_f1']:.4f}, AUC: {auc_str}")
    
    # Create analysis files
    main_logger.info("Creating Excel analysis...")
    excel_path = create_excel_mcc_analysis(all_results, output_dir)
    
    main_logger.info("Generating comparative visualizations...")
    best_model = plot_comparative_results(all_results, output_dir)
    
    # Copy best model
    best_model_source = os.path.join(
        output_dir, 
        f"lr_{best_model['hyperparameters']['learning_rate']}_bs_{best_model['hyperparameters']['batch_size']}", 
        "best_model.pkl"
    )
    best_model_dest = os.path.join(output_dir, "best_model.pkl")
    shutil.copy2(best_model_source, best_model_dest)
    
    # Save all results
    with open(os.path.join(output_dir, 'all_results.json'), 'w') as f:
        json.dump(all_results, f, indent=4, default=lambda x: float(x) if isinstance(x, np.float32) else x)
    
    # Final summary
    mcc_scores = [result['test_metrics']['mcc'] for result in all_results]
    auc_str = f"{best_model['test_metrics']['auc']:.4f}" if best_model['test_metrics']['auc'] else 'N/A'
    
    print("\n" + "="*70)
    print("MCC-Focused Hyperparameter Tuning Completed Successfully!")
    print(f"Results saved to: {output_dir}")
    print(f"Best model: lr={best_model['hyperparameters']['learning_rate']}, "
          f"bs={best_model['hyperparameters']['batch_size']}")
    print(f"Best model metrics:")
    print(f"  MCC (PRIMARY): {best_model['test_metrics']['mcc']:.4f}")
    print(f"  Accuracy: {best_model['test_metrics']['accuracy']:.4f}")
    print(f"  F1 Score: {best_model['test_metrics']['pos_f1']:.4f}")
    print(f"  AUC: {auc_str}")
    print(f"  TP: {best_model['test_metrics']['tp']}, TN: {best_model['test_metrics']['tn']}, "
          f"FP: {best_model['test_metrics']['fp']}, FN: {best_model['test_metrics']['fn']}")
    print(f"\nMCC Statistics:")
    print(f"  Mean MCC: {np.mean(mcc_scores):.4f}")
    print(f"  Best MCC: {max(mcc_scores):.4f}")
    print(f"  Experiments with MCC > 0: {sum(1 for mcc in mcc_scores if mcc > 0)}/{len(mcc_scores)}")
    print("="*70)


if __name__ == "__main__":
    main()